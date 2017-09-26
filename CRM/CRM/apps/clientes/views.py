#-*- coding: utf-8 -*-
from django.shortcuts import render
from django.shortcuts import render_to_response
from django.template import RequestContext
from CRM.apps.clientes.forms import *
from django.contrib.auth import login, logout, authenticate
from django.http import HttpResponseRedirect
from django.contrib.auth.models import User ##
from CRM.apps.clientes.models import *
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, InvalidPage#paginator
from django.http.response import HttpResponse
from django.core.mail import EmailMultiAlternatives
from django.contrib import messages
from openpyxl import Workbook
from django.conf import settings
from django.core.files import File
import os 

def index_view(request): 
	
	return render(request,'clientes/index.html')

def login_view(request):
	mensaje = ""
	if request.user.is_authenticated():#verificacmos si el usuario ya esta authenticado o logueado
		return HttpResponseRedirect('/index/')#si esta logueado lo redirigimos a la pagina principal
	else: #si no esta authenticado 
		if request.method == "POST":
			formulario = Login_form(request.POST) #creamos un objeto de Loguin_form
			if formulario.is_valid(): #si la informacion enviada es correcta		
				usu= formulario.cleaned_data['usuario'] #guarda informacion ingresada del formulario
				pas= formulario.cleaned_data['clave'] #guarda informacion ingresada del formulario
				usuario = authenticate(username = usu,password = pas)#asigna la autenticacion del usuario
				if usuario is not None and usuario.is_active:#si el usuario no es nulo y esta activo
					login(request,usuario)#se loguea al sistema con la informacion de usuario
					return HttpResponseRedirect('/index/')#redirigimos a la pagina principal
				else:
					mensaje = "usuario y/o clave incorrecta"
		formulario = Login_form() #creamos un formulario nuevo limpio
		ctx = {'form':formulario, 'mensaje':mensaje}#variable de contexto para pasar info a login.html
		return render_to_response('clientes/login.html',ctx, context_instance = RequestContext(request))

def logout_view(request):
	logout(request)# funcion de django importda anteriormente
	return HttpResponseRedirect('/')# redirigimos a la pagina principal
##################################### USUARIOS ######################################################

def user_view(request): 
	us = User.objects.get(id= request.user.id)
	ctx={'user':us}
	return render_to_response('clientes/user.html',ctx,context_instance = RequestContext(request))

def edit_user_view(request):
	info = ""	
	info_enviado = False
	us = User.objects.get(id = request.user.id)
	user = User_profile.objects.get(user=us)
	#formulario = UserForm()
	if request.method == "POST":
		formulario = UserForm(request.POST, request.FILES, instance = us)
		#form_user= User_profile_Form(request.POST, request.FILES, instance = user)		
		if formulario.is_valid():
			info_enviado = True
			#telefono = formulario.cleaned_data['telefono']
			#clave = formulario.cleaned_data['password']
			#formulario.password = us.set_password(clave)
			edit_user = formulario.save(commit = False)
			#edit_user_telefono = form_user.save(commit = False)
			#edit_user_telefono.telefono= telefono
			#formulario.save_m2m()
			#edit_user.status = True
			#edit_user_telefono.save()
			edit_user.save()
			info = "Guardado Satisfactoriamente"
			#return HttpResponseRedirect ('/')
			#return HttpResponseRedirect ('/user/')
	else:
		formulario = UserForm(instance = us)
		#form_user = User_profile_Form(instance = user)
	ctx = {'form':formulario,'informacion':info, 'info_enviado':info_enviado}	
	return render_to_response('clientes/edit_user.html',ctx,context_instance = RequestContext(request))

def edit_password_view(request):
	info = ""	
	info_enviado = False
	us = User.objects.get(id = request.user.id)
	if request.method == "POST":
		formulario = PasswordForm(request.POST, request.FILES, instance = us)		
		if formulario.is_valid():
			info_enviado = True
			clave = formulario.cleaned_data['password']
			usu= us.username #guarda informacion ingresada del formulario
			pas= clave #guarda informacion ingresada del formulario
			formulario.password = us.set_password(clave)
			edit_user = formulario.save(commit = False)			
			#formulario.save_m2m()
			#edit_user.status = True
			edit_user.save()
			info = "Guardado Satisfactoriamente"
			usuario = authenticate(username = usu,password = pas)#asigna la autenticacion del usuario
			if usuario is not None and usuario.is_active:#si el usuario no es nulo y esta activo
				login(request,usuario)#se loguea al sistema con la informacion de usuario
				return HttpResponseRedirect('/password_guardado_user/')#redirigimos a la pagina principal
			#return HttpResponseRedirect ('/')
			#return HttpResponseRedirect ('/index/')
	else:
		formulario = PasswordForm(instance = us)
	ctx = {'form':formulario, 'informacion':info, 'info_enviado':info_enviado}	
	return render_to_response('clientes/edit_password.html',ctx,context_instance = RequestContext(request))

def password_guardado_user_view(request):
	
	return render(request, 'clientes/password_guardado_user.html') 	


####################################################################################################
#
##################################### PACIENTES ####################################################


def registrar_paciente_view(request):
	usuario = User.objects.get(id= request.user.id)
	form = Register_Persona_Form()
	form_paciente = Register_Persona_Paciente_Form()
	if request.method == "POST":
		form = Register_Persona_Form(request.POST)
		form_paciente = Register_Persona_Paciente_Form(request.POST)
		if form.is_valid() and form_paciente.is_valid():
			#nombres = form.cleaned_data['first_name']
			#apellidos = form.cleaned_data['last_name']
			#usuario = form.cleaned_data['username']
			#email = form.cleaned_data['email']
			#password_one = form.cleaned_data['password_one']
			#password_two = form.cleaned_data['password_two']
			#telefono = form.cleaned_data['telefono']
			#supervisor  = form.cleaned_data['supervisor']
			#u = User.objects.create_user(first_name=nombres,last_name=apellidos,username=usuario,email=email,password=password_one)
			persona = form.save(commit=False)
			persona.empresa=usuario.user_profile.empresa
			persona.tipo="Paciente" 
			persona.save()
			persona_paciente = form_paciente.save(commit=False)
			persona_paciente.persona = persona	
			persona_paciente.save()		
			#user= form.save(commit=False)
			#user.user_profile.telefono=telefono
			#user.save()
			#user = u.save(commit=False)
			#user.user_profile.telefono=telefono
			
			ctx = {'paciente':persona}
			return render_to_response('clientes/thanks_register.html',ctx,context_instance=RequestContext(request))
		else:		
			ctx = {'form':form,'form_paciente':form_paciente}
			return render_to_response('clientes/registrar_paciente.html',ctx,context_instance=RequestContext(request))
	ctx = {'form':form,'form_paciente':form_paciente}
	return render_to_response('clientes/registrar_paciente.html',ctx,context_instance=RequestContext(request))

def pacientes_view(request,pagina):
	usuario = User.objects.get(id= request.user.id)
	lista_pacientes = Persona.objects.filter(tipo="Paciente",empresa=usuario.user_profile.empresa)	
	primera = "<<Primera"
	ultima = "Ultima>>"	
	query = request.GET.get('q','')     
	if query:
		qset = (
			Q(nombres__icontains=query)|
			Q(apellidos__icontains=query)|
			Q(cedula__icontains=query)|
			Q(correo__icontains=query)
		)
		results = Persona.objects.filter(qset,tipo="Paciente",empresa=usuario.user_profile.empresa).distinct()  
		mostrar = False      
	else:
		mostrar = True
		results = []

	#lista_prod = Producto.objects.filter(status = True)#SELECT * from Producto WHERE status= True
	paginator = Paginator(lista_pacientes, 2) 
	try:
		page = int(pagina)
	except:
		page = 1
	try:
		pacientes = paginator.page(page)
	except (EmptyPage,InvalidPage):
		pacientes = paginator.page(paginator.num_pages)	


	return render_to_response('clientes/pacientes.html', {
		"results": results,
		"query": query,
		"mostrar": mostrar,
		"pacientes":pacientes,
		"lista_pacientes":lista_pacientes, 
		"primera":primera,
		"ultima":ultima,       
	},context_instance=RequestContext(request))	

def ver_paciente_view(request, id_paciente): 
	paciente = Persona.objects.get(id= id_paciente)
	ctx={'paciente':paciente}
	return render(request,'clientes/ver_paciente.html',ctx)		

def editar_paciente_view(request,id_paciente):
	info = ""	
	info_enviado = False
	persona = Persona.objects.get(id = id_paciente)
	paciente = Paciente.objects.get(persona=persona)
	if request.method == "POST":
		formulario = Register_Persona_Form(request.POST, request.FILES, instance = persona)	
		form_paciente= Register_Persona_Paciente_Form(request.POST, request.FILES, instance = paciente)		
		if formulario.is_valid() and form_paciente.is_valid():
			info_enviado = True
			#clave = formulario.cleaned_data['password']
			#formulario.password = us.set_password(clave)
			edit_persona = formulario.save(commit = False)	
			edit_persona_paciente = form_paciente.save(commit = False)		
			#edit_user.status = True
			edit_persona_paciente.save()
			edit_persona.save()
			info = "Guardado Satisfactoriamente"
			#return HttpResponseRedirect ('/')
			#return HttpResponseRedirect ('/instructores/')
	else:
		formulario = Register_Persona_Form(instance = persona)
		form_paciente = Register_Persona_Paciente_Form(instance = paciente)
	ctx = {'form':formulario, 'form_paciente':form_paciente,'informacion':info, 'info_enviado':info_enviado,'paciente':persona}	
	return render_to_response('clientes/editar_paciente.html',ctx,context_instance = RequestContext(request))			

def eliminar_paciente_view(request,id_paciente):	
	persona = Persona.objects.get(id= id_paciente)
	ctx={'paciente':persona}
	return render(request, 'clientes/eliminar_paciente.html',ctx)

def confirmar_eliminar_paciente_view(request,id_paciente):	
	
	info = "inicializando"
	try:
		persona = Persona.objects.get(id= id_paciente)
		persona_paciente = Paciente.objects.get(persona = persona)
		persona_interesado = Interesado.objects.get(persona = persona)
		
		persona.delete()			
		persona_paciente.delete()
		if persona_interesado:
			persona_interesado.delete()
		
		info = "El paciente ha sido eliminado correctamente"
		return HttpResponseRedirect('/pacientes/page/1')
	except:
		info = "EL paciente no se puede eliminar"
		#return render_to_response('home/productos.html', context_instance = RequestContext(request))
		return HttpResponseRedirect('/index/')		

######################################################################################################

##################################### INTERESADOS ####################################################

def registrar_interesado_view(request):
	usuario = User.objects.get(id= request.user.id)
	form = Register_Persona_Form()
	form_interesado = Register_Persona_Interesado_Form()
	if request.method == "POST":
		form = Register_Persona_Form(request.POST)
		form_interesado = Register_Persona_Interesado_Form(request.POST)
		if form.is_valid() and form_interesado.is_valid():
			#nombres = form.cleaned_data['first_name']
			#apellidos = form.cleaned_data['last_name']
			#usuario = form.cleaned_data['username']
			#email = form.cleaned_data['email']
			#password_one = form.cleaned_data['password_one']
			#password_two = form.cleaned_data['password_two']
			#telefono = form.cleaned_data['telefono']
			#supervisor  = form.cleaned_data['supervisor']
			#u = User.objects.create_user(first_name=nombres,last_name=apellidos,username=usuario,email=email,password=password_one)
			interesado = form.save(commit=False)
			interesado.empresa=usuario.user_profile.empresa
			interesado.tipo="Interesado" 
			interesado.save()
			persona_interesado = form_interesado.save(commit=False)
			persona_interesado.persona = interesado	
			persona_interesado.save()		
			#user= form.save(commit=False)
			#user.user_profile.telefono=telefono
			#user.save()
			#user = u.save(commit=False)
			#user.user_profile.telefono=telefono
			
			ctx = {'interesado':interesado}
			return render_to_response('clientes/thanks_register_interesado.html',ctx,context_instance=RequestContext(request))
		else:		
			ctx = {'form':form,'form_interesado':form_interesado}
			return render_to_response('clientes/registrar_interesado.html',ctx,context_instance=RequestContext(request))
	ctx = {'form':form,'form_interesado':form_interesado}
	return render_to_response('clientes/registrar_interesado.html',ctx,context_instance=RequestContext(request))

def interesados_view(request,pagina):
	usuario = User.objects.get(id= request.user.id)
	lista_interesados = Persona.objects.filter(tipo="Interesado",empresa=usuario.user_profile.empresa)	
	primera = "<<Primera"
	ultima = "Ultima>>"	
	query = request.GET.get('q','')     
	if query:
		qset = (
			Q(nombres__icontains=query)|
			Q(apellidos__icontains=query)|
			Q(cedula__icontains=query)|
			Q(correo__icontains=query)
		)
		results = Persona.objects.filter(qset,tipo="Interesado",empresa=usuario.user_profile.empresa).distinct()  
		mostrar = False      
	else:
		mostrar = True
		results = []

	#lista_prod = Producto.objects.filter(status = True)#SELECT * from Producto WHERE status= True
	paginator = Paginator(lista_interesados, 2) 
	try:
		page = int(pagina)
	except:
		page = 1
	try:
		interesados = paginator.page(page)
	except (EmptyPage,InvalidPage):
		interesados = paginator.page(paginator.num_pages)	


	return render_to_response('clientes/interesados.html', {
		"results": results,
		"query": query,
		"mostrar": mostrar,
		"interesados":interesados,
		"lista_interesados":lista_interesados, 
		"primera":primera,
		"ultima":ultima,       
	},context_instance=RequestContext(request))	

def ver_interesado_view(request, id_interesado):
 	info = ""	
	info_enviado = False	
	persona = Persona.objects.get(id = id_interesado)
	interesado = Interesado.objects.get(persona=persona)
	if request.method == "POST":
		formulario = Interesado_Form(request.POST, request.FILES, instance = interesado)	
		
		if formulario.is_valid():
			info_enviado = True
			#clave = formulario.cleaned_data['password']
			#formulario.password = us.set_password(clave)
			edit_interesado = formulario.save(commit = False)	
			
			edit_interesado.save()
			info = "Guardado Satisfactoriamente"
			#return HttpResponseRedirect ('/')
			#return HttpResponseRedirect ('/instructores/')
	else:
		formulario = Interesado_Form(instance = interesado)		

	ctx={'interesado':persona,'form':formulario, 'info_enviado':info_enviado,'informacion':info}
	return render(request,'clientes/ver_interesado.html',ctx)	


def editar_interesado_view(request,id_interesado):
	info = ""	
	info_enviado = False
	persona = Persona.objects.get(id = id_interesado)
	interesado = Interesado.objects.get(persona=persona)
	if request.method == "POST":
		formulario = Register_Persona_Form(request.POST, request.FILES, instance = persona)	
		form_interesado= Register_Persona_Interesado_Form(request.POST, request.FILES, instance = interesado)		
		if formulario.is_valid() and form_interesado.is_valid():
			info_enviado = True
			#clave = formulario.cleaned_data['password']
			#formulario.password = us.set_password(clave)
			edit_persona = formulario.save(commit = False)	
			edit_persona_interesado = form_interesado.save(commit = False)		
			#edit_user.status = True
			edit_persona_interesado.save()
			edit_persona.save()
			info = "Guardado Satisfactoriamente"
			#return HttpResponseRedirect ('/')
			#return HttpResponseRedirect ('/instructores/')
	else:
		formulario = Register_Persona_Form(instance = persona)
		form_interesado = Register_Persona_Interesado_Form(instance = interesado)
	ctx = {'form':formulario, 'form_interesado':form_interesado,'informacion':info, 'info_enviado':info_enviado,'interesado':persona}	
	return render_to_response('clientes/editar_interesado.html',ctx,context_instance = RequestContext(request))			

def eliminar_interesado_view(request,id_interesado):	
	persona = Persona.objects.get(id= id_interesado)
	ctx={'interesado':persona}
	return render(request, 'clientes/eliminar_interesado.html',ctx)

def confirmar_eliminar_interesado_view(request,id_interesado):	
	
	info = "inicializando"
	try:
		persona = Persona.objects.get(id= id_interesado)
		persona_interesado = Interesado.objects.get(persona = persona)
		
		
		persona.delete()			
		persona_interesado.delete()
		
		
		info = "El interesado ha sido eliminado correctamente"
		return HttpResponseRedirect('/interesados/page/1')
	except:
		info = "EL interesado no se puede eliminar"
		#return render_to_response('home/productos.html', context_instance = RequestContext(request))
		return HttpResponseRedirect('/index/')	

def interesado_a_paciente_view(request,id_interesado):	
	#usuario = User.objects.get(id= request.user.id)
	per = Persona.objects.get(id = id_interesado)	
	#form = Interesado_a_Paciente_Form()
	form_paciente = Register_Persona_Paciente_Form()
	if request.method == "POST":
		#form = Interesado_a_Paciente_Form(request.POST, request.FILES, instance = per)		
		form_paciente = Register_Persona_Paciente_Form(request.POST)
		if form_paciente.is_valid():
			#nombres = form.cleaned_data['first_name']
			#apellidos = form.cleaned_data['last_name']
			#usuario = form.cleaned_data['username']
			#email = form.cleaned_data['email']
			#password_one = form.cleaned_data['password_one']
			#password_two = form.cleaned_data['password_two']
			#telefono = form.cleaned_data['telefono']
			#supervisor  = form.cleaned_data['supervisor']
			#u = User.objects.create_user(first_name=nombres,last_name=apellidos,username=usuario,email=email,password=password_one)
			#persona = form.save(commit=False)
			per.tipo="Paciente"
			per.save()
			persona_paciente = form_paciente.save(commit=False)
			persona_paciente.persona = per	
			persona_paciente.save()		
			#user= form.save(commit=False)
			#user.user_profile.telefono=telefono
			#user.save()
			#user = u.save(commit=False)
			#user.user_profile.telefono=telefono
			
			ctx = {'paciente':per}
			return render_to_response('clientes/thanks_register.html',ctx,context_instance=RequestContext(request))
		else:		
			ctx = {'form_paciente':form_paciente,'paciente':per}
			return render_to_response('clientes/interesado_a_paciente.html',ctx,context_instance=RequestContext(request))
	ctx = {'form_paciente':form_paciente,'paciente':per}
	return render_to_response('clientes/interesado_a_paciente.html',ctx,context_instance=RequestContext(request))

######################################################################################################

########################################### REPORTES #################################################

def reporte_exel_pacientes_view(request):	
	user = User.objects.get(id= request.user.id)
	lista_consultar = Persona.objects.filter(tipo="Paciente",empresa=user.user_profile.empresa)

	wb = Workbook()
	#Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
	ws = wb.active
	#En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
	#ws['B1'] = 'REPORTE DE PACIENTES '
	#Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
	#ws.merge_cells('B1:E1')
	#Creamos los encabezados desde la celda B3 hasta la E3
	ws['B1'] = 'NOMBRES'
	ws['C1'] = 'APELLIDOS'
	ws['D1'] = 'CEDULA'
	ws['E1'] = 'CORREO' 
	ws['F1'] = 'MOVIL' 
	ws['G1'] = 'TELEFONO'
	ws['H1'] = 'DEPARTAMENTO'
	ws['I1'] = 'CIUDAD'
	ws['J1'] = 'DIRECCION'
	ws['K1'] = 'GENERO'
	ws['L1'] = 'GRUPO SANGUINEO'
	ws['M1'] = 'FECHA NACIMIENTO'      
	cont=2
	#Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
	for persona in lista_consultar:
		ws.cell(row=cont,column=2).value = persona.nombres
		ws.cell(row=cont,column=3).value = persona.apellidos
		ws.cell(row=cont,column=4).value = persona.cedula
		ws.cell(row=cont,column=5).value = persona.correo
		ws.cell(row=cont,column=6).value = persona.movil
		ws.cell(row=cont,column=7).value = persona.telefono
		ws.cell(row=cont,column=8).value = persona.departamento
		ws.cell(row=cont,column=9).value = persona.ciudad
		ws.cell(row=cont,column=10).value = persona.direccion
		ws.cell(row=cont,column=11).value = persona.paciente.genero
		ws.cell(row=cont,column=12).value = persona.paciente.grupo_sanguineo
		ws.cell(row=cont,column=13).value = persona.paciente.fecha_nacimiento
		cont = cont + 1
	#Establecemos el nombre del archivo
	nombre_archivo ="ReportePacientes.xlsx"
	#Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
	response = HttpResponse(content_type="application/ms-excel") 
	contenido = "attachment; filename={0}".format(nombre_archivo)
	response["Content-Disposition"] = contenido
	wb.save(response)
	return response					

def reporte_exel_interesados_view(request):	
	user = User.objects.get(id= request.user.id)
	lista_consultar = Persona.objects.filter(tipo="Interesado",empresa=user.user_profile.empresa)

	wb = Workbook()
	#Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
	ws = wb.active
	#En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
	#ws['B1'] = 'REPORTE DE PACIENTES '
	#Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
	#ws.merge_cells('B1:E1')
	#Creamos los encabezados desde la celda B3 hasta la E3
	ws['B1'] = 'NOMBRES'
	ws['C1'] = 'APELLIDOS'
	ws['D1'] = 'CEDULA'
	ws['E1'] = 'CORREO' 
	ws['F1'] = 'MOVIL' 
	ws['G1'] = 'TELEFONO'
	ws['H1'] = 'DEPARTAMENTO'
	ws['I1'] = 'CIUDAD'
	ws['J1'] = 'DIRECCION'
	ws['K1'] = 'PROCEDENCIA'
	ws['L1'] = 'NIVEL INTERES'
	     
	cont=2
	#Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
	for persona in lista_consultar:
		ws.cell(row=cont,column=2).value = persona.nombres
		ws.cell(row=cont,column=3).value = persona.apellidos
		ws.cell(row=cont,column=4).value = persona.cedula
		ws.cell(row=cont,column=5).value = persona.correo
		ws.cell(row=cont,column=6).value = persona.movil
		ws.cell(row=cont,column=7).value = persona.telefono
		ws.cell(row=cont,column=8).value = persona.departamento
		ws.cell(row=cont,column=9).value = persona.ciudad
		ws.cell(row=cont,column=10).value = persona.direccion
		ws.cell(row=cont,column=11).value = persona.interesado.procedencia
		ws.cell(row=cont,column=12).value = persona.interesado.nivel_interes
		cont = cont + 1
	#Establecemos el nombre del archivo
	nombre_archivo ="ReporteInteresados.xlsx"
	#Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
	response = HttpResponse(content_type="application/ms-excel") 
	contenido = "attachment; filename={0}".format(nombre_archivo)
	response["Content-Disposition"] = contenido
	wb.save(response)
	return response						